from bs4 import BeautifulSoup
import os
from openpyxl import Workbook, load_workbook

path = r"C:\Users\Win10\Desktop\apps\utilities\total-express-shopee\nfs"
os.chdir(path)

dataEmissao = input("Qual a data de emissão? (Ex: 25/03/2022): ")
arqXls = input("Cole o nome do arquivo xls aqui: ")
arqXls = arqXls + '.xlsx'
cont = 2

def read_text_file(file_path,dataEmissao,cont,arqXls):
    with open(file_path, 'r') as f:
        data = f.read() #lê um arquivo N

        Bs_data = BeautifulSoup(data, features="xml")

        b_num = Bs_data.find('nNF')  # num da nf
        numSerial = "001"  # num de série
        b_numAces = Bs_data.find('chNFe')  # num de acesso
        b_vTotal = Bs_data.find('vNF')  # valor total da nota
        b_vProd = Bs_data.find('vProd')  # valor total dos produtos
        b_numCfop = Bs_data.find('CFOP')  # num CFOP

        numNF = str(b_num.string)  # num nf to string
        numAces = str(b_numAces.string)  # num de acesso to string
        vTotal = str(b_vTotal.string)  # valor total da nota to string
        vProd = str(b_vProd.string)  # valor total dos produtos to string
        numCfop = str(b_numCfop.string)  # num CFOP to string

        wb = load_workbook(arqXls) #abre o arquivo xls
        ws = wb.active #dxar a aba ativa pra mexer
        ws["B" + str(cont)] = "00000" + numNF #preencher num nf
        ws["C" + str(cont)] = numSerial  # preencher num de série
        ws["D" + str(cont)] = numAces  # preencher num de acesso
        ws["E" + str(cont)] = dataEmissao  # preencher data de emissão
        ws["F" + str(cont)] = vTotal.replace('.',',')  # preencher valor total
        ws["G" + str(cont)] = vProd.replace('.',',')  # preencher valor produtos
        ws["H" + str(cont)] = numCfop  # preencher CFOP
        wb.save(arqXls) #salvar arq modificado

        print("------------------------{}----------------------------------".format(cont))
        print("Número da NF= Valor: {} / Tipo: {} / Tamanho: {}".format(numNF, type(numNF), len(numNF)))
        print("Número de Série= Valor: {} / Tipo: {} / Tamanho: {}".format(numSerial, type(numSerial), len(numSerial)))
        print("Chave de Acesso: {} / Tipo: {} / Tamanho: {}".format(numAces, type(numAces), len(numAces)))
        print("Data de Emissão: {} / Tipo: {} / Tamanho: {}".format(dataEmissao, type(dataEmissao), len(dataEmissao)))
        print("Valor Total da Nota: {} / Tipo: {} / Tamanho: {}".format(vTotal, type(vTotal), len(vTotal)))
        print("Valor Total dos Produtos: {} / Tipo: {} / Tamanho: {}".format(vProd, type(vProd), len(vProd)))
        print("CFOP: {} / Tipo: {} / Tamanho: {}".format(numCfop, type(numCfop), len(numCfop)))
        print("--------------------------------------------------------")

for file in os.listdir():

    if file.endswith(".xml"):
        file_path = f"{path}\{file}"
        read_text_file(file_path, dataEmissao,cont,arqXls)
        cont+=1
